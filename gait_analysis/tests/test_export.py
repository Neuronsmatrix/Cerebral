import openpyxl
from matplotlib.figure import Figure

from modules.visualization.export import export_figure_png, export_results_csv, export_results_xlsx

RESULTS = {
    "spatiotemporal": {"cadence_steps_per_min": 110.0, "speed_m_per_s": 1.2},
    "joint_angles_mean": {"left_knee": [0.0, 30.0, 60.0]},
}


def test_export_csv_roundtrip(tmp_path):
    p = tmp_path / "r.csv"
    export_results_csv(RESULTS, p)
    text = p.read_text()
    assert "cadence_steps_per_min" in text and "110.0" in text


def test_export_xlsx_has_expected_sheets(tmp_path):
    p = tmp_path / "r.xlsx"
    export_results_xlsx(RESULTS, p)
    wb = openpyxl.load_workbook(p)
    assert "spatiotemporal" in wb.sheetnames
    assert "joint_angles_mean" in wb.sheetnames


def test_export_png_creates_nonempty_file(tmp_path):
    fig = Figure()
    fig.add_subplot(111).plot([0, 1], [0, 1])
    p = tmp_path / "f.png"
    export_figure_png(fig, p)
    assert p.exists() and p.stat().st_size > 0
