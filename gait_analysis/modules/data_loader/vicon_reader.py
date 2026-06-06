"""Read Vicon XLSX into the unified pose DataFrame (real export schema)."""
import numpy as np
import pandas as pd
from openpyxl import load_workbook


def _read_rows(path: str, max_rows: int | None = None) -> list[list]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for i, r in enumerate(ws.iter_rows(values_only=True)):
        rows.append(list(r))
        if max_rows is not None and i + 1 >= max_rows:
            break
    wb.close()
    return rows


def _find_axis_row(rows: list[list]) -> int:
    """Index of the row that contains the X/Y/Z axis labels.

    The real axis row also holds 'Frame'/'Sub Frame', so we look for a row where
    X/Y/Z dominate rather than one that is exclusively X/Y/Z.
    """
    for i, row in enumerate(rows):
        cells = [str(c).strip().upper() if c is not None else "" for c in row]
        xyz = sum(c in ("X", "Y", "Z") for c in cells)
        if xyz >= 3:
            return i
    raise ValueError("Could not locate the X/Y/Z axis row in Vicon XLSX")


def _marker_name(cell) -> str | None:
    """Clean marker token: strip subject prefix ('Subject:LKNE' -> 'LKNE')."""
    if cell is None or not str(cell).strip():
        return None
    name = str(cell).strip()
    if ":" in name:
        name = name.split(":", 1)[1].strip()
    return name


def load_vicon_xlsx(filepath: str, vicon_fps: float = 100.0) -> pd.DataFrame:
    """Parse a Vicon Trajectories export into the unified pose DataFrame.

    - axis row auto-detected (tolerant of Frame/Sub Frame columns);
    - marker names are colon-split to drop the subject prefix;
    - the duplicate '| ... |' block yields names like 'LKNE |' that simply
      become extra columns and are dropped at mapping time;
    - mm->m when median(|coord|) > 10;
    - adds zero-based ``timestamp`` = (frame - frame[0]) / vicon_fps and ``frame``.
    """
    rows = _read_rows(filepath)
    axis_i = _find_axis_row(rows)
    name_row = rows[axis_i - 1]
    axis_row = rows[axis_i]

    # forward-fill marker names across the 3 (X,Y,Z) columns they span
    names: list[str | None] = []
    last = None
    for c in name_row:
        nm = _marker_name(c)
        if nm is not None:
            last = nm
        names.append(last)

    # locate the Frame column in the axis row (first cell whose label is 'Frame')
    frame_col = None
    for j, a in enumerate(axis_row):
        if a is not None and str(a).strip().lower() == "frame":
            frame_col = j
            break

    columns: list[str | None] = []
    for name, axis in zip(names, axis_row):
        ax = str(axis).strip().upper() if axis is not None else ""
        if name is None or ax not in ("X", "Y", "Z"):
            columns.append(None)
            continue
        columns.append(f"{name}_{ax.lower()}")

    data_rows = rows[axis_i + 1:]
    records, frames = [], []
    for r in data_rows:
        if all(c is None for c in r):
            continue
        rec = {}
        for col, val in zip(columns, r):
            if col is None or val is None:
                continue
            try:
                rec[col] = float(val)
            except (TypeError, ValueError):
                continue
        if not rec:
            continue
        records.append(rec)
        fv = r[frame_col] if frame_col is not None and frame_col < len(r) else None
        try:
            frames.append(float(fv))
        except (TypeError, ValueError):
            frames.append(float(len(frames)))

    df = pd.DataFrame(records)
    # drop the units row (e.g. all 'mm') which produced an empty/NaN record set:
    df = df.dropna(how="all").reset_index(drop=True)

    coord = df.to_numpy(dtype=float)
    if np.nanmedian(np.abs(coord)) > 10.0:
        df = df / 1000.0

    frames = np.asarray(frames[: len(df)], dtype=float)
    df.insert(0, "timestamp", (frames - frames[0]) / vicon_fps)
    df.insert(0, "frame", range(len(df)))
    return df


def map_vicon_to_caliscope(vicon_df: pd.DataFrame, mapping: dict) -> pd.DataFrame:
    """Rename ``MARKER_{x,y,z}`` columns to ``canonical_{x,y,z}`` via ``mapping``."""
    rename = {}
    for marker, landmark in mapping.items():
        for axis in ("x", "y", "z"):
            rename[f"{marker}_{axis}"] = f"{landmark}_{axis}"
    return vicon_df.rename(columns=rename)
